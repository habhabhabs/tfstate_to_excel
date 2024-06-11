import json
import openpyxl
import argparse
import os
import time
from openpyxl.utils import get_column_letter

def safe_str(value):
    """Convert a value to a string, handling non-string types gracefully."""
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return str(value)

def safe_sheet_title(title):
    """Ensure the sheet title is within Excel's 31-character limit."""
    return title[:31]

def main(tfstate_path, include_data_sources, output_path):
    start_time = time.time()

    try:
        # Load the Terraform state file
        with open(tfstate_path, 'r') as file:
            state = json.load(file)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        summary_ws = wb.active
        summary_ws.title = 'Summary'

        # Dictionary to hold worksheets by resource type
        sheets = {}

        # Dictionary to track all unique attribute keys per resource type
        attribute_keys = {}

        # Dictionary to track the number of records per resource type
        record_counts = {}

        # First pass: Collect all unique attribute keys and count records
        for resource in state['resources']:
            try:
                # Filter to include only resources (exclude data sources)
                if resource.get('mode') != 'managed' and not include_data_sources:
                    continue

                resource_type = resource['type']
                
                if resource_type not in attribute_keys:
                    attribute_keys[resource_type] = set()
                    record_counts[resource_type] = 0
                
                for instance in resource.get('instances', []):
                    attributes = instance.get('attributes', {})
                    attribute_keys[resource_type].update(attributes.keys())
                    record_counts[resource_type] += 1
            except KeyError as e:
                print(f"Key error while processing resource {resource}: {e}")
            except Exception as e:
                print(f"Unexpected error while processing resource {resource}: {e}")

        # Second pass: Create sheets and write data
        for resource in state['resources']:
            try:
                # Filter to include only resources (exclude data sources)
                if resource.get('mode') != 'managed' and not include_data_sources:
                    continue

                resource_type = resource['type']
                resource_name = resource['name']
                
                # Get or create a worksheet for the resource type
                sheet_title = safe_sheet_title(resource_type)
                if sheet_title not in sheets:
                    ws = wb.create_sheet(title=sheet_title)
                    sheets[sheet_title] = ws
                    # Define the headers for the new sheet
                    headers = ['Resource Name', 'Resource Address'] + sorted(attribute_keys[resource_type])
                    ws.append(headers)
                else:
                    ws = sheets[sheet_title]
                
                # Iterate through instances within the resource
                for instance in resource.get('instances', []):
                    resource_address = instance.get('address', 'N/A')
                    attributes = instance.get('attributes', {})
                    
                    row = [resource_name, resource_address]
                    for key in sorted(attribute_keys[resource_type]):
                        row.append(safe_str(attributes.get(key, '')))
                    
                    ws.append(row)
            except KeyError as e:
                print(f"Key error while processing resource {resource}: {e}")
            except Exception as e:
                print(f"Unexpected error while processing resource {resource}: {e}")

        # Add summary information
        summary_ws.append(['Resource Type', 'Record Count'])
        for resource_type, count in record_counts.items():
            summary_ws.append([resource_type, count])

        # Add provider information
        providers = set(resource['provider'] for resource in state['resources'] if 'provider' in resource)
        summary_ws.append([])
        summary_ws.append(['Providers'])
        for provider in providers:
            summary_ws.append([provider])

        # Add account information (assuming account_name and account_number are fields in the state)
        account_name = state.get('account_name', 'N/A')
        account_number = state.get('account_number', 'N/A')
        summary_ws.append([])
        summary_ws.append(['Account Name', account_name])
        summary_ws.append(['Account Number', account_number])

        # Adjust column widths for each sheet
        for ws in sheets.values():
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # Adjust column widths for the summary sheet
        for col in summary_ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            summary_ws.column_dimensions[column].width = adjusted_width

        # Save the workbook
        wb.save(output_path)

        end_time = time.time()
        execution_time = end_time - start_time

        print("\nExecution Summary:")
        print(f"Source File: {os.path.abspath(tfstate_path)}")
        print(f"Output File: {os.path.abspath(output_path)}")
        print(f"Execution Time: {execution_time:.2f} seconds")

    except FileNotFoundError:
        print(f"The '{tfstate_path}' file was not found. Please ensure the file is present in the specified directory.")
    except json.JSONDecodeError:
        print("Failed to decode JSON. Please ensure the 'terraform.tfstate' file is a valid JSON.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='(2024) Alex KM - MIT License - Extract Terraform state to an Excel workbook.')
    parser.add_argument('--tfstate', required=True, help='Path to the Terraform state (.tfstate) file.')
    parser.add_argument('--include-data-sources', action='store_true', help='Include data sources in the extraction.')
    parser.add_argument('--output', default='terraform_state.xlsx', help='Path to the output Excel file (default: terraform_state.xlsx).')

    args = parser.parse_args()

    # Ensure the output path has .xlsx suffix
    if not args.output.endswith('.xlsx'):
        print("Error: The output file must have a .xlsx suffix.")
    else:
        main(args.tfstate, args.include_data_sources, args.output)
